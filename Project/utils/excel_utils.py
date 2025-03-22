# utils/excel_utils.py

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from tkinter import simpledialog

def sort_detections(detections, margin_value):
    # Sort detections by y1 (grouping values within the margin value)
    sorted_by_y1 = sorted(detections, key=lambda x: x[1])

    # Group y1 values within the margin value
    grouped_detections = []
    current_group = []
    current_y1 = None

    for det in sorted_by_y1:
        if current_y1 is None or abs(det[1] - current_y1) <= margin_value:
            current_group.append(det)
        else:
            # Sort the current group by x1
            current_group.sort(key=lambda x: x[0])
            grouped_detections.append(current_group)
            current_group = [det]
        current_y1 = det[1]

    # Add the last group
    if current_group:
        current_group.sort(key=lambda x: x[0])
        grouped_detections.append(current_group)

    return grouped_detections

def save_results_to_csv(detections, margin_value):
    csv_filename = simpledialog.askstring("Save CSV", "Enter the name for the CSV file (without extension):")
    if not csv_filename:
        print("CSV filename not provided. Exiting...")
        return

    # Sort and group detections
    grouped_detections = sort_detections(detections, margin_value)

    # Create an Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Detections"

    # Write the header row
    headers = ["Hole No.", "x1", "y1", "x2", "y2"]
    ws.append(headers)

    # Assign hole numbers in the desired format (line, column)
    line_number = 1
    for group in grouped_detections:
        column_number = 1
        for det in group:
            hole_number = f"{line_number},{column_number}"
            ws.append([hole_number, det[0], det[1], det[2], det[3]])
            column_number += 1
        line_number += 1

    # Apply table formatting
    format_excel_table(ws)

    # Save the workbook
    excel_file_path = f"{csv_filename}.xlsx"
    wb.save(excel_file_path)
    print(f"Excel file saved as {excel_file_path}")

def format_excel_table(ws):
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define font style for headers
    header_font = Font(bold=True)

    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # Format header row
    for cell in ws[1]:
        cell.font = header_font

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name (e.g., 'A', 'B')
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width