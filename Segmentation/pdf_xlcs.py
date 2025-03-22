import tkinter as tk
from tkinter import filedialog, simpledialog
import cv2
import numpy as np
from ultralytics import YOLO
from fpdf import FPDF
import os
import random
import csv
from tkinter import simpledialog
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

# Global variables
conf_threshold = 0.5
scale_factor = 0.5

def get_parameters():
    global conf_threshold, scale_factor

    def apply_settings():
        global conf_threshold, scale_factor
        conf_threshold = conf_scale.get() / 100
        scale_factor = scale_conv.get() / 100
        param_window.destroy()

    def exit_program():
        print("Leaving.")
        param_window.destroy()
        exit()

    # Tkinter window for parameter settings
    param_window = tk.Tk()
    param_window.title("Configurations")

    # Confidence level slider
    tk.Label(param_window, text="Trust Factor").pack()
    conf_scale = tk.Scale(param_window, from_=1, to=100, orient=tk.HORIZONTAL)
    conf_scale.set(50)
    conf_scale.pack()

    # Scale factor slider
    tk.Label(param_window, text="Scale").pack()
    scale_conv = tk.Scale(param_window, from_=1, to=100, orient=tk.HORIZONTAL)
    scale_conv.set(50)
    scale_conv.pack()

    # Apply button
    apply_button = tk.Button(param_window, text="Apply", command=apply_settings)
    apply_button.pack()

    # Exit button
    exit_button = tk.Button(param_window, text="Exit", command=exit_program)
    exit_button.pack()

    param_window.mainloop()


# Function to draw circle perimeters around detections
def draw_circle_perimeters(frame, detections):
    for x1, y1, x2, y2 in detections:
        center = (int((x1 + x2) / 2), int((y1 + y2) / 2))
        radius = int(max((x2 - x1), (y2 - y1)) / 2)
        cv2.circle(frame, center, radius, (255, 255, 255), 1, cv2.LINE_AA)


# Function to create the zoomed overlay with measurements
def draw_measurements_overlay(frame, bbox, scale_factor):
    x1, y1, x2, y2 = bbox
    center_x, center_y = (x1 + x2) // 2, (y1 + y2) // 2
    radius = int(max((x2 - x1), (y2 - y1)) / 2)  # Radius of the bounding box in pixels

    pixel_to_cm = 0.3125 / 10 * scale_factor
    dist_top_cm = radius * pixel_to_cm
    dist_right_cm = radius * pixel_to_cm

    fixed_radius = 200
    overlay = np.zeros((640, 640, 3), dtype=np.uint8)
    overlay.fill(50)

    canvas_center_x, canvas_center_y = 320, 320
    cv2.circle(overlay, (canvas_center_x, canvas_center_y), fixed_radius, (255, 255, 255), 2, cv2.LINE_AA)
    cv2.arrowedLine(overlay, (canvas_center_x, canvas_center_y), (canvas_center_x, canvas_center_y - fixed_radius),
                    (0, 0, 255), 2, tipLength=0.1)
    cv2.arrowedLine(overlay, (canvas_center_x, canvas_center_y), (canvas_center_x + fixed_radius, canvas_center_y),
                    (0, 255, 0), 2, tipLength=0.1)

    font_scale = 0.6
    font_thickness = 2
    cv2.putText(overlay, f"{dist_top_cm:.2f} cm", (canvas_center_x - 80, canvas_center_y - fixed_radius - 20),
                cv2.FONT_HERSHEY_SIMPLEX, font_scale, (0, 0, 255), font_thickness)
    cv2.putText(overlay, f"{dist_right_cm:.2f} cm", (canvas_center_x + fixed_radius + 10, canvas_center_y),
                cv2.FONT_HERSHEY_SIMPLEX, font_scale, (0, 255, 0), font_thickness)

    return overlay


# Function to save results as a PDF with text above images and centered overlay
def save_results_as_pdf(original_image, segmented_image, hole_overlay, detections, conf_threshold, scale_factor):
    pdf_filename = simpledialog.askstring("Save PDF", "Enter the name for the PDF file (without extension):")
    if not pdf_filename:
        print("PDF filename not provided. Exiting...")
        return

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, "Detection Results", ln=True, align="C")

    # Adding the parameters
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, f"Trust Factor: {conf_threshold*100:.2f}%", ln=True, align="C")
    pdf.cell(200, 10, f"Scale Factor: {scale_factor*100:.2f}%", ln=True, align="C")
    pdf.ln(10)

    # Save temporary images for the PDF
    original_path = "original_temp.jpg"
    segmented_path = "segmented_temp.jpg"
    overlay_path = "overlay_temp.jpg"
    cv2.imwrite(original_path, original_image)
    cv2.imwrite(segmented_path, segmented_image)
    cv2.imwrite(overlay_path, hole_overlay)

    # Add 'Original' text above the original image
    pdf.set_font("Arial", "B", 12)
    pdf.cell(200, 10, "Original", ln=True, align="C")
    pdf.image(original_path, x=70, w=100)
    pdf.ln(10)

    # Add 'Segmented' text above the segmented image
    pdf.cell(200, 10, "Segmented", ln=True, align="C")
    pdf.image(segmented_path, x=70, w=100)
    pdf.ln(10)

    # Add 'Dimensions' text above the overlay image
    pdf.cell(200, 10, "Dimensions", ln=True, align="C")
    pdf.image(overlay_path, x=70, w=100)
    pdf.ln(10)

    pdf.output(f"{pdf_filename}.pdf")

    # Clean up temporary files
    os.remove(original_path)
    os.remove(segmented_path)
    os.remove(overlay_path)
    print(f"PDF saved as {pdf_filename}.pdf")


def sort_detections(detections):
    # Sort detections by y1 (grouping values within a margin of 20)
    sorted_by_y1 = sorted(detections, key=lambda x: x[1])

    # Group y1 values within a margin of 20
    grouped_detections = []
    current_group = []
    current_y1 = None

    for det in sorted_by_y1:
        if current_y1 is None or abs(det[1] - current_y1) <= 20:
            current_group.append(det)
        else:
            # Sort the current group by x1
            current_group.sort(key=lambda x: x[0])
            grouped_detections.append(current_group)
            current_group = [det]
        current_y1 = det[1]

    # Add the last group
    if current_group:
        current_group.sort(key=lambda x: x[0])
        grouped_detections.append(current_group)

    return grouped_detections

def save_results_to_csv(detections):
    csv_filename = simpledialog.askstring("Save CSV", "Enter the name for the CSV file (without extension):")
    if not csv_filename:
        print("CSV filename not provided. Exiting...")
        return

    # Sort and group detections
    grouped_detections = sort_detections(detections)

    # Create an Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Detections"

    # Write the header row
    headers = ["Hole No.", "x1", "y1", "x2", "y2"]
    ws.append(headers)

    # Assign hole numbers in the desired format (line, column)
    line_number = 1
    for group in grouped_detections:
        column_number = 1
        for det in group:
            hole_number = f"{line_number},{column_number}"
            ws.append([hole_number, det[0], det[1], det[2], det[3]])
            column_number += 1
        line_number += 1

    # Apply table formatting
    format_excel_table(ws)

    # Save the workbook
    excel_file_path = f"{csv_filename}.xlsx"
    wb.save(excel_file_path)
    print(f"Excel file saved as {excel_file_path}")

def format_excel_table(ws):
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define font style for headers
    header_font = Font(bold=True)

    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # Format header row
    for cell in ws[1]:
        cell.font = header_font

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name (e.g., 'A', 'B')
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

# Load the YOLO model
model = YOLO("C:/Users/diogo/Desktop/python/yolov11/yolov11s/modelo3.pt")

# Main program logic
while True:
    get_parameters()  # Ask for parameters

    image_path = filedialog.askopenfilename(title="Select an image", filetypes=[("Images", "*.jpg;*.png;*.jpeg")])
    if not image_path:
        print("Cancel.")
        break

    frame = cv2.imread(image_path)
    if frame is None:
        print("Error.")
        continue

    frame_resized = cv2.resize(frame, (640, 640))
    original_image = frame_resized.copy()

    results = model.predict(frame_resized, conf=conf_threshold, verbose=False, max_det=1000)
    detections = [(int(box.xyxy[0][0]), int(box.xyxy[0][1]), int(box.xyxy[0][2]), int(box.xyxy[0][3])) for box in
                  results[0].boxes]

    # Randomly select a bounding box for overlay visualization
    if detections:
        selected_bbox = random.choice(detections)
        hole_overlay = draw_measurements_overlay(original_image, selected_bbox, scale_factor)

    # Create segmented image with circle perimeters
    segmented_image = np.zeros_like(frame_resized)
    draw_circle_perimeters(segmented_image, detections)

    # Save results as PDF and CSV
    save_results_as_pdf(original_image, segmented_image, hole_overlay, detections, conf_threshold, scale_factor)
    save_results_to_csv(detections)

    break
